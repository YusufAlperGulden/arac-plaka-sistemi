import codecs
import csv
import io
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from report_exports import REPORT_COLUMNS, export_csv, export_pdf, export_xlsx


SAMPLE_RECORD = {
    "status": "completed",
    "action_type": "Kurum İçi Operasyonlar",
    "add_date": "30.07.2026 09:00:00",
    "vehicle_name": "RENAULT 2016 CLIO",
    "plate": "34EZS794",
    "driver": "Seda ŞAHİN",
    "request_no": "TAL-2026-15",
    "service_form_no": "SRV-88",
    "start_mileage": "151800",
    "end_mileage": "151821",
    "start_date": "30.07.2026 09:05:00",
    "distance": "21",
    "end_date": "30.07.2026 10:00:00",
    "notes": "Araç sorunsuz teslim edildi.",
}


class ReportExportTests(unittest.TestCase):
    def test_csv_has_utf8_bom_headers_and_turkish_content(self):
        payload = export_csv([SAMPLE_RECORD])

        self.assertTrue(payload.startswith(codecs.BOM_UTF8))
        rows = list(
            csv.reader(io.StringIO(payload.decode("utf-8-sig"), newline=""))
        )
        self.assertEqual(tuple(rows[0]), REPORT_COLUMNS)
        self.assertEqual(rows[1][0], "Tamamlandı")
        self.assertEqual(rows[1][3], "RENAULT 2016 CLIO")
        self.assertEqual(rows[1][5], "Seda ŞAHİN")
        self.assertEqual(rows[1][13], "Araç sorunsuz teslim edildi.")

    def test_csv_guards_spreadsheet_formula_prefixes(self):
        record = {
            **SAMPLE_RECORD,
            "driver": "+cmd|' /C calc'!A0",
            "request_no": "-1+2",
            "service_form_no": "@SUM(A1:A2)",
            "notes": " =HYPERLINK(\"https://example.invalid\")",
        }

        rows = list(
            csv.reader(
                io.StringIO(
                    export_csv([record]).decode("utf-8-sig"),
                    newline="",
                )
            )
        )

        self.assertEqual(rows[1][5], "'+cmd|' /C calc'!A0")
        self.assertEqual(rows[1][6], "'-1+2")
        self.assertEqual(rows[1][7], "'@SUM(A1:A2)")
        self.assertEqual(
            rows[1][13],
            "' =HYPERLINK(\"https://example.invalid\")",
        )

    def test_xlsx_has_zip_magic_headers_content_and_formula_guards(self):
        record = {
            **SAMPLE_RECORD,
            "driver": "=2+3",
            "request_no": "+SUM(A1:A2)",
            "service_form_no": "-2+3",
            "notes": "@malicious",
        }

        payload = export_xlsx([record])

        self.assertTrue(payload.startswith(b"PK\x03\x04"))
        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        self.addCleanup(workbook.close)
        worksheet = workbook["Araç Raporu"]
        self.assertEqual(
            tuple(cell.value for cell in worksheet[1]),
            REPORT_COLUMNS,
        )
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.cell(2, 1).value, "Tamamlandı")
        self.assertEqual(worksheet.cell(2, 5).value, "34EZS794")
        self.assertEqual(worksheet.cell(2, 6).value, "'=2+3")
        self.assertEqual(worksheet.cell(2, 7).value, "'+SUM(A1:A2)")
        self.assertEqual(worksheet.cell(2, 8).value, "'-2+3")
        self.assertEqual(worksheet.cell(2, 14).value, "'@malicious")
        for column in (6, 7, 8, 14):
            self.assertEqual(worksheet.cell(2, column).data_type, "s")

    def test_empty_exports_keep_the_common_headers(self):
        csv_rows = list(
            csv.reader(
                io.StringIO(export_csv([]).decode("utf-8-sig"), newline="")
            )
        )
        self.assertEqual(csv_rows, [list(REPORT_COLUMNS)])

        workbook = load_workbook(
            io.BytesIO(export_xlsx([])),
            read_only=True,
        )
        self.addCleanup(workbook.close)
        worksheet = workbook["Araç Raporu"]
        self.assertEqual(
            tuple(cell.value for cell in next(worksheet.iter_rows())),
            REPORT_COLUMNS,
        )

    def test_pdf_has_magic_page_markers_and_eof(self):
        payload = export_pdf([SAMPLE_RECORD])

        self.assertTrue(payload.startswith(b"%PDF-"))
        self.assertIn(b"/Type /Page", payload)
        self.assertTrue(payload.rstrip().endswith(b"%%EOF"))
        self.assertGreater(len(payload), 2_000)

    def test_pdf_ascii_fallback_keeps_basic_content_safe(self):
        with patch(
            "report_exports._resolve_pdf_font",
            return_value=("Helvetica", False),
        ):
            payload = export_pdf([SAMPLE_RECORD])

        self.assertIn(b"Arac Hareket Raporu", payload)
        self.assertIn(b"RENAULT 2016 CLIO", payload)
        self.assertIn(b"34EZS794", payload)
        self.assertIn(b"Seda SAHIN", payload)


if __name__ == "__main__":
    unittest.main()
