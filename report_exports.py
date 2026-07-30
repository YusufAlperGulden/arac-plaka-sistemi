"""Framework-independent helpers for exporting vehicle movement reports."""

from __future__ import annotations

import csv
import io
import unicodedata
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape


REPORT_COLUMNS = (
    "Durum",
    "Hareket Tipi",
    "Eklenme Tarihi",
    "Araç",
    "Plaka",
    "Sürücü",
    "Talep No",
    "Servis Formu",
    "İlk Sayaç",
    "Son Sayaç",
    "Başlangıç",
    "Yapılan Mesafe",
    "Bitiş",
    "Açıklama",
)

_FIELD_ALIASES = {
    "Hareket Tipi": ("action_type", "movement_type"),
    "Eklenme Tarihi": ("add_date", "created_at"),
    "Araç": ("vehicle_name", "vehicle_label", "vehicle"),
    "Plaka": ("plate",),
    "Sürücü": ("driver", "user"),
    "Talep No": ("request_no",),
    "Servis Formu": ("service_form_no",),
    "İlk Sayaç": ("start_mileage",),
    "Son Sayaç": ("end_mileage",),
    "Başlangıç": ("start_date", "started_at"),
    "Yapılan Mesafe": ("distance",),
    "Bitiş": ("end_date", "ended_at"),
    "Açıklama": ("notes", "description"),
}

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_PDF_FONT_CANDIDATES = (
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("C:/Windows/Fonts/arial.ttf"),
)
_PDF_UNICODE_FONT_NAME = "VehicleReportUnicode"
_ASCII_TRANSLATION = str.maketrans(
    {
        "Ç": "C",
        "Ğ": "G",
        "İ": "I",
        "Ö": "O",
        "Ş": "S",
        "Ü": "U",
        "ç": "c",
        "ğ": "g",
        "ı": "i",
        "ö": "o",
        "ş": "s",
        "ü": "u",
    }
)


def _get_field(record: Any, key: str) -> Any:
    if isinstance(record, Mapping):
        return record.get(key)
    return getattr(record, key, None)


def _first_field(record: Any, keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = _get_field(record, key)
        if value is not None:
            return value
    return None


def _status_value(record: Any) -> Any:
    status = _get_field(record, "status")
    if status not in (None, ""):
        normalized = str(status).strip().lower()
        if normalized in {"active", "open", "aktif"}:
            return "Aktif"
        if normalized in {"completed", "closed", "tamamlandı", "tamamlandi"}:
            return "Tamamlandı"
        return status

    is_active = _get_field(record, "is_active")
    if is_active is not None:
        return "Aktif" if bool(is_active) else "Tamamlandı"

    end_value = _first_field(record, ("end_date", "ended_at"))
    return "Aktif" if end_value in (None, "") else "Tamamlandı"


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _spreadsheet_safe(value: Any) -> str:
    """Return text that spreadsheet programs cannot interpret as a formula."""
    text = _stringify(value)
    significant = text.lstrip(" \t\r\n")
    if significant.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _row_values(record: Any, *, spreadsheet_safe: bool) -> list[str]:
    values: list[Any] = []
    for heading in REPORT_COLUMNS:
        if heading == "Durum":
            value = _status_value(record)
        else:
            value = _first_field(record, _FIELD_ALIASES[heading])
        values.append(
            _spreadsheet_safe(value)
            if spreadsheet_safe
            else _stringify(value)
        )
    return values


def _materialize_records(records: Iterable[Any]) -> list[Any]:
    if records is None:
        return []
    return list(records)


def export_csv(records: Iterable[Any]) -> bytes:
    """Export records as an Excel-friendly UTF-8 CSV with a BOM."""
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(REPORT_COLUMNS)
    for record in _materialize_records(records):
        writer.writerow(_row_values(record, spreadsheet_safe=True))
    return output.getvalue().encode("utf-8-sig")


def export_xlsx(records: Iterable[Any]) -> bytes:
    """Export records as a styled XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _materialize_records(records)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Araç Raporu"
    worksheet.freeze_panes = "A2"
    worksheet.append(REPORT_COLUMNS)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in rows:
        worksheet.append(_row_values(record, spreadsheet_safe=True))

    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        maximum_length = max(
            len(_stringify(cell.value))
            for cell in column_cells
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(maximum_length + 2, 10),
            36,
        )
        for cell in column_cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _ascii_safe(value: str) -> str:
    translated = value.translate(_ASCII_TRANSLATION)
    return (
        unicodedata.normalize("NFKD", translated)
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def _resolve_pdf_font() -> tuple[str, bool]:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    if _PDF_UNICODE_FONT_NAME in registered_fonts:
        return _PDF_UNICODE_FONT_NAME, True

    for candidate in _PDF_FONT_CANDIDATES:
        if not candidate.is_file():
            continue
        try:
            pdfmetrics.registerFont(
                TTFont(_PDF_UNICODE_FONT_NAME, str(candidate))
            )
            return _PDF_UNICODE_FONT_NAME, True
        except Exception:
            continue

    return "Helvetica", False


def _pdf_paragraph(value: str, style: Any, *, unicode_font: bool) -> Any:
    from reportlab.platypus import Paragraph

    text = value if unicode_font else _ascii_safe(value)
    safe_markup = escape(text).replace("\n", "<br/>")
    return Paragraph(safe_markup or " ", style)


def export_pdf(records: Iterable[Any]) -> bytes:
    """Export records as a landscape PDF table."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

    rows = _materialize_records(records)
    font_name, unicode_font = _resolve_pdf_font()
    output = io.BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=7 * mm,
        bottomMargin=11 * mm,
        title="Araç Hareket Raporu",
        author="Araç Plaka Sistemi",
        pageCompression=0,
    )

    title_style = ParagraphStyle(
        "ReportTitle",
        fontName=font_name,
        fontSize=12,
        leading=14,
        alignment=TA_CENTER,
        spaceAfter=4 * mm,
    )
    header_style = ParagraphStyle(
        "ReportHeader",
        fontName=font_name,
        fontSize=5.3,
        leading=6.1,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        "ReportCell",
        fontName=font_name,
        fontSize=5.2,
        leading=6.2,
    )

    title = _pdf_paragraph(
        "Araç Hareket Raporu",
        title_style,
        unicode_font=unicode_font,
    )
    table_data = [
        [
            _pdf_paragraph(
                heading,
                header_style,
                unicode_font=unicode_font,
            )
            for heading in REPORT_COLUMNS
        ]
    ]
    for record in rows:
        table_data.append(
            [
                _pdf_paragraph(value, cell_style, unicode_font=unicode_font)
                for value in _row_values(record, spreadsheet_safe=False)
            ]
        )

    width_weights = (7, 10, 9, 13, 8, 9, 8, 8, 7, 7, 9, 8, 9, 13)
    total_weight = sum(width_weights)
    column_widths = [
        document.width * weight / total_weight
        for weight in width_weights
    ]
    table = Table(
        table_data,
        colWidths=column_widths,
        repeatRows=1,
        hAlign="LEFT",
    )
    table_style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for row_number in range(1, len(table_data)):
        if row_number % 2 == 0:
            table_style_commands.append(
                (
                    "BACKGROUND",
                    (0, row_number),
                    (-1, row_number),
                    colors.HexColor("#F1F5F9"),
                )
            )
    table.setStyle(TableStyle(table_style_commands))

    generated_text = datetime.now().strftime("%d.%m.%Y %H:%M")

    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 7)
        footer_y = 5 * mm
        canvas.setFillColor(colors.HexColor("#475569"))
        canvas.drawString(
            document.leftMargin,
            footer_y,
            (
                f"Oluşturulma: {generated_text}"
                if unicode_font
                else f"Olusturulma: {generated_text}"
            ),
        )
        canvas.drawRightString(
            page_size[0] - document.rightMargin,
            footer_y,
            (
                f"Sayfa {doc.page}"
                if unicode_font
                else f"Sayfa {doc.page}"
            ),
        )
        canvas.restoreState()

    document.build(
        [title, Spacer(1, 2 * mm), table],
        onFirstPage=draw_footer,
        onLaterPages=draw_footer,
    )
    return output.getvalue()
